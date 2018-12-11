import datetime
import decimal
import hashlib
import operator

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate
from django.contrib.auth.models import Group, Permission
from django.contrib.messages.views import SuccessMessageMixin
from django.core.exceptions import ObjectDoesNotExist, PermissionDenied
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Q
from django.http import Http404
from django.shortcuts import HttpResponse, redirect, render
from django.urls import reverse
from django.utils.encoding import force_text
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import FormView, View
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from borgia.utils import (GroupLateralMenuMixin, GroupPermissionMixin,
                          UserMixin, shop_from_group)
from events.forms import (EventAddWeightForm, EventCreateForm,
                            EventDeleteForm, EventDownloadXlsxForm,
                            EventFinishForm, EventListForm,
                            EventSelfRegistrationForm,
                            EventUpdateForm, EventUploadXlsxForm)
from events.models import Event


class EventList(GroupPermissionMixin, FormView,
                      GroupLateralMenuMixin):
    template_name = 'finances/event_list.html'
    lm_active = 'lm_event_list'
    perm_codename = 'view_event'
    form_class = EventListForm

    def dispatch(self, request, *args, **kwargs):
        try:
            self.group = Group.objects.get(name=kwargs['group_name'])
        except ObjectDoesNotExist:
            raise Http404
        # For safety, but it shouldn't be possible with regex pattern in url FOR NOW (If Post is used, url need to be modified)
        except ValueError:
            raise Http404

        return super(EventList, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(EventList, self).get_context_data(**kwargs)
        shared_events = Event.objects.filter(
            date__gte=datetime.date.today(), done=False).order_by('-date')

        for se in shared_events:  # Duplicate
            # Si fini, on recupere la participation, sinon la preinscription
            se.weight_of_user = se.get_weight_of_user(
                self.request.user, se.done)
            se.number_registrants = se.get_number_registrants()
            se.number_participants = se.get_number_participants()
            se.total_weights_registrants = se.get_total_weights_registrants()
            se.total_weights_participants = se.get_total_weights_participants()
            try:
                se.has_perm_manage = (self.request.user == se.manager or
                                      Permission.objects.get(codename='change_event') in self.group.permissions.all())
            except ObjectDoesNotExist:
                se.has_perm_manage = False
        context['shared_events'] = shared_events
        # Permission SelfRegistration
        try:
            if Permission.objects.get(codename='self_register_event') in self.group.permissions.all():
                context['has_perm_self_register_event'] = True
        except ObjectDoesNotExist:
            pass

        return context

    def form_valid(self, form, **kwargs):

        date_begin = form.cleaned_data['date_begin']
        date_end = form.cleaned_data['date_end']
        order_by = form.cleaned_data['order_by']
        done = form.cleaned_data['done']

        if date_begin is None:
            date_begin = datetime.date.min
        if date_end is None:
            date_end = datetime.date.max

        if done != "both":
            if done == "done":
                done = True
            else:
                done = False

            shared_events = Event.objects.filter(
                date__range=[date_begin, date_end], done=done)
        else:
            shared_events = Event.objects.filter(
                date__range=[date_begin, date_end])

        context = self.get_context_data(**kwargs)
        if order_by != '-date':
            shared_events = shared_events.order_by(order_by).order_by('-date')
        else:
            shared_events = shared_events.order_by('-date')

        for se in shared_events:  # Duplicate
            # Si fini, on recupere la participation, sinon la preinscription
            se.weight_of_user = se.get_weight_of_user(
                self.request.user, se.done)
            se.number_registrants = se.get_number_registrants()
            se.number_participants = se.get_number_participants()
            se.total_weights_registrants = se.get_total_weights_registrants()
            se.total_weights_participants = se.get_total_weights_participants()

        context['shared_events'] = shared_events

        return self.render_to_response(context)


class EventCreate(GroupPermissionMixin, SuccessMessageMixin, FormView,
                        GroupLateralMenuMixin):
    form_class = EventCreateForm
    template_name = 'finances/event_create.html'
    success_url = None
    perm_codename = 'add_event'
    lm_active = 'lm_event_create'
    success_message = "'%(description)s' a bien été créé."

    def form_valid(self, form):

        se = Event.objects.create(
            description=form.cleaned_data['description'],
            date=form.cleaned_data['date'],
            allow_self_registeration=form.cleaned_data['allow_self_registeration'],
            manager=self.request.user)
        if form.cleaned_data['price']:
            se.price = decimal.Decimal(form.cleaned_data['price'])
        if form.cleaned_data['bills']:
            se.bills = form.cleaned_data['bills']
        if form.cleaned_data['date_end_registration']:
            date_end_registration = form.cleaned_data['date_end_registration']
            if date_end_registration > datetime.date.today() and date_end_registration <= form.cleaned_data['date']:
                se.date_end_registration = form.cleaned_data['date_end_registration']

        se.save()
        self.se = se

        return super(EventCreate, self).form_valid(form)

    def get_success_message(self, cleaned_data):
        return self.success_message % dict(
            description=self.se.description,
        )

    def get_success_url(self):
        return reverse('url_event_update',
                       kwargs={'group_name': self.group.name, 'pk': self.se.pk}
                       )


class EventDelete(GroupPermissionMixin, SuccessMessageMixin, FormView, GroupLateralMenuMixin):
    """
    Delete a event and redirect to the list of events.

    :param kwargs['group_name']: name of the group used.
    :param kwargs['pk']: pk of the event
    :param self.perm_codename: codename of the permission checked.
    """
    template_name = 'finances/event_delete.html'
    form_class = EventDeleteForm
    success_url = None
    perm_codename = None  # Checked in dispatch
    success_message = "L'évènement a bien été supprimé."

    def dispatch(self, request, *args, **kwargs):
        try:
            self.group = Group.objects.get(name=kwargs['group_name'])
            self.se = Event.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404

        # Permissions
        try:
            if Permission.objects.get(codename='change_event') not in self.group.permissions.all():
                if request.user != self.se.manager:
                    raise Http404
        except ObjectDoesNotExist:
            raise Http404
        if self.se.done:
            raise PermissionDenied

        return super(EventDelete, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(EventDelete, self).get_context_data(**kwargs)
        context['se'] = self.se
        return context

    def form_valid(self, form):
        self.se.delete()
        return super(EventDelete, self).form_valid(form)

    def get_success_url(self):
        return reverse('url_event_list',
                       kwargs={'group_name': self.group.name}
                       )


class EventFinish(GroupPermissionMixin, SuccessMessageMixin, FormView, GroupLateralMenuMixin):
    """
    Finish a event and redirect to the list of events.
    This command is used when you want to keep the event in the database, but
    you don't want to pay in Borgia (for instance paid with real money).

    :param kwargs['group_name']: name of the group used.
    :param kwargs['pk']: pk of the event
    :param self.perm_codename: codename of the permission checked.
    """
    form_class = EventFinishForm
    template_name = 'finances/event_finish.html'
    success_url = None
    perm_codename = None  # Checked in dispatch
    success_message = "'%(description)s' a bien été terminé."

    def dispatch(self, request, *args, **kwargs):
        try:
            self.se = Event.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404

        # Permissions
        try:
            group = Group.objects.get(name=kwargs['group_name'])
            if Permission.objects.get(codename='proceed_payment_event') not in group.permissions.all():
                raise Http404
        except ObjectDoesNotExist:
            raise Http404

        if self.se.done:
            raise PermissionDenied

        # Check if there are participants
        self.total_weights_participants = self.se.get_total_weights_participants()
        if self.total_weights_participants == 0:
            return redirect(reverse(
                'url_event_update',
                kwargs={'group_name': group.name, 'pk': self.se.pk}
            ) + '?no_participant=True')

        # Get some data :
        self.price = self.se.price
        try:
            self.ponderation_price = self.price / self.total_weights_participants
        except TypeError:
            self.ponderation_price = 0

        return super(EventFinish, self).dispatch(request, *args, **kwargs)

    def get_initial(self):
        """
        Populate the form with the current price.
        """
        initial = super(EventFinish, self).get_initial()
        initial['total_price'] = self.price
        initial['ponderation_price'] = round(self.ponderation_price, 2)
        return initial

    def get_context_data(self, **kwargs):
        context = super(EventFinish, self).get_context_data(**kwargs)
        context['se'] = self.se
        context['total_price'] = self.se.price
        context['total_weights_participants'] = self.total_weights_participants
        context['ponderation_price'] = round(self.ponderation_price, 2)
        return context

    def form_valid(self, form):
        type_payment = form.cleaned_data['type_payment']
        if (type_payment == 'pay_by_total'):
            self.se.pay_by_total(self.request.user, User.objects.get(
                pk=1), form.cleaned_data['total_price'])

        if (type_payment == 'pay_by_ponderation'):
            self.se.pay_by_ponderation(self.request.user, User.objects.get(
                pk=1), form.cleaned_data['ponderation_price'])

        if (type_payment == 'no_payment'):
            self.se.end_without_payment(form.cleaned_data['remark'])

        return super(EventFinish, self).form_valid(form)

    def get_success_message(self, cleaned_data):
        return self.success_message % dict(
            description=self.se.description,
        )

    def get_success_url(self):
        return reverse('url_event_list',
                       kwargs={'group_name': self.group.name}
                       )


class EventUpdate(GroupPermissionMixin, SuccessMessageMixin, FormView, GroupLateralMenuMixin):
    """
    Update the Shared Event,
    """
    form_class = EventUpdateForm
    template_name = 'finances/event_update.html'
    perm_codename = None  # Checked in dispatch()
    success_message = "'%(description)s' a bien été mis à jour."

    manager_changed = False

    def dispatch(self, request, *args, **kwargs):
        try:
            self.group = Group.objects.get(name=kwargs['group_name'])
            self.se = Event.objects.get(pk=int(kwargs['pk']))
        except ObjectDoesNotExist:
            raise Http404
        # For safety, but it shouldn't be possible with regex pattern in url FOR NOW (If Post is used, url need to be modified)
        except ValueError:
            raise Http404

        # Permissions
        try:
            if Permission.objects.get(codename='change_event') not in self.group.permissions.all():
                if request.user != self.se.manager:
                    raise Http404
        except ObjectDoesNotExist:
            raise Http404

        self.no_participant = request.GET.get(
            'no_participant', False) == "True"  # "True" == "True"

        return super(EventUpdate, self).dispatch(request, *args, **kwargs)

    def get_initial(self):
        initial = super(EventUpdate, self).get_initial()
        initial['price'] = self.se.price
        initial['bills'] = self.se.bills
        initial['manager'] = self.se.manager.username
        initial['allow_self_registeration'] = self.se.allow_self_registeration
        return initial

    def get_context_data(self, **kwargs):

        # list_year() contains smth like [2011, 2015, ...]

        context = super(EventUpdate, self).get_context_data(**kwargs)
        context['pk'] = self.se.pk
        context['done'] = self.se.done
        if self.se.done:
            context['remark'] = self.se.remark
            context['price'] = self.se.price
        # Pour les users
        context['number_registrants'] = self.se.get_number_registrants
        context['number_participants'] = self.se.get_number_participants
        context['total_weights_registrants'] = self.se.get_total_weights_registrants
        context['total_weights_participants'] = self.se.get_total_weights_participants
        context['no_participant'] = self.no_participant

        # Création des forms excel
        context['upload_xlsx_form'] = EventUploadXlsxForm()
        context['download_xlsx_form'] = EventDownloadXlsxForm()

        # Permission FinishEvent
        try:
            group = Group.objects.get(name=context['group_name'])
            if Permission.objects.get(codename='proceed_payment_event') in group.permissions.all():
                context['has_perm_proceed_payment'] = True
        except ObjectDoesNotExist:
            pass  # has_perm not True

        return context

    def form_valid(self, form):
        if form.cleaned_data['price']:
            self.se.price = form.cleaned_data['price']
        if form.cleaned_data['bills']:
            self.se.bills = form.cleaned_data['bills']
        if form.cleaned_data['manager']:
            form_manager = User.objects.get(
                username=form.cleaned_data['manager'])
            manage_permission = False
            if self.se.manager != form_manager:
                for group in form_manager.groups.all():
                    if Permission.objects.get(codename='add_event') in group.permissions.all():
                        manage_permission = True
                        break
                if manage_permission:
                    self.se.manager = form_manager
                    self.manager_changed = True
                else:
                    messages.warning(self.request,
                                     "%(user)s ne dispose pas de droits suffisants pour gérer l'évènement" % dict(
                                         user=form_manager))
        self.se.allow_self_registeration = form.cleaned_data['allow_self_registeration']
        self.se.save()

        return super(EventUpdate, self).form_valid(form)

    def get_success_message(self, cleaned_data):
        if self.manager_changed:
            return "%(user)s gère désormais l'évènement" % dict(
                user=self.se.manager,
            )
        else:
            return self.success_message % dict(
                description=self.se.description,
            )

    def get_success_url(self):
        if self.manager_changed and Permission.objects.get(codename='change_event') not in self.group.permissions.all():
            return reverse('url_group_workboard',
                           kwargs={'group_name': self.group.name}
                           )
        else:
            return reverse('url_event_update',
                           kwargs={'group_name': self.group.name,
                                   'pk': self.se.pk}
                           )


class EventSelfRegistration(GroupPermissionMixin, SuccessMessageMixin, FormView, GroupLateralMenuMixin):
    """
    Allow a user to register himself

    :param kwargs['group_name']: name of the group used.
    :param kwargs['pk']: pk of the event
    :param self.perm_codename: codename of the permission checked.
    """
    form_class = EventSelfRegistrationForm
    template_name = 'finances/event_self_registration.html'
    perm_codename = 'self_register_event'

    def dispatch(self, request, *args, **kwargs):
        try:
            self.se = Event.objects.get(pk=int(kwargs['pk']))
        except ObjectDoesNotExist:
            raise Http404
        # For safety, but it shouldn't be possible with regex pattern in url FOR NOW (If Post is used, url need to be modified)
        except ValueError:
            raise Http404

        # Permissions
        if self.se.done:
            raise PermissionDenied

        if not self.se.allow_self_registeration:
            raise PermissionDenied

        if self.se.date_end_registration:
            if datetime.date.today() > self.se.date_end_registration:
                raise PermissionDenied

        return super(EventSelfRegistration, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(EventSelfRegistration,
                        self).get_context_data(**kwargs)
        context['shared_event'] = self.se
        context['registeration_of_user'] = self.se.get_weight_of_user(
            self.request.user, False)  # Duplicate
        return context

    def get_initial(self):
        initial = super(EventSelfRegistration, self).get_initial()
        initial['weight'] = self.se.get_weight_of_user(
            self.request.user, False)  # Duplicate
        return initial

    def form_valid(self, form):
        self.new_weight = int(form.cleaned_data['weight'])
        self.se.change_weight(self.request.user, self.new_weight, False)
        return super(EventSelfRegistration, self).form_valid(form)

    def get_success_message(self, cleaned_data):
        if self.new_weight > 0:
            return "Vous avez bien été inscrit"
        else:
            return "Vous avez bien été désinscrit"

    def get_success_url(self):
        return reverse('url_event_self_registration',
                       kwargs={'group_name': self.group.name, 'pk': self.se.pk}
                       )


class EventChangeWeight(GroupPermissionMixin, View):
    perm_codename = None

    def get(self, request, *args, **kwargs):
        """
        Change la valeur de la pondération d'un participant user pour un événement
        Permissions :   Si événements terminé -> denied,
                        Si pas manager ou pas la perm 'finances.change_event' -> denied
        :param pk: pk de l'événement
        :param user_pk: paramètre GET correspondant au pk de l'user
        :param pond_pk: paramètre GET correspondant à la nouvelle pondération
        :type pk, user_pk, pond_pk: int
        """
        response = 0

        try:
            # Variables d'entrées
            se = Event.objects.get(pk=kwargs['pk'])
            user = User.objects.get(pk=kwargs['user_pk'])
            pond = int(request.GET['pond'])
            isParticipant = int(request.GET['isParticipant'])  # Boolean

            # Permission
            if request.user != se.manager and request.user.has_perm('finances.change_event') is False:
                raise PermissionDenied
            # Même en ayant la permission, on ne modifie plus une event terminé
            elif se.done is True:
                raise PermissionDenied

            if pond > 0:
                if isParticipant in [0, 1]:
                    # Changement de la pondération
                    se.change_weight(user, pond, isParticipant)

                    # Réponse
                    response = 1

        except KeyError:
            pass
        except ObjectDoesNotExist:
            pass
        except ValueError:
            pass

        return HttpResponse(response)


class EventManageUsers(GroupPermissionMixin, FormView, GroupLateralMenuMixin):
    """
    Manage the users. The get displays the list of concerned users. The post form add weight to one of them.
    """
    form_class = EventAddWeightForm
    template_name = 'finances/event_manage_users.html'
    perm_codename = None  # Checked in dispatch()

    def get_list_weights(self, state):
        se = Event.objects.get(pk=self.kwargs['pk'])
        if state == 'users':
            return se.list_users_weight()
        elif state == 'participants':
            return se.list_participants_weight()
        elif state == 'registrants':
            return se.list_registrants_weight()

    def dispatch(self, request, *args, **kwargs):
        try:
            self.group = Group.objects.get(name=kwargs['group_name'])
            self.se = Event.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404

        # Permission
        if request.user != self.se.manager or not request.user.has_perm('finances.change_event'):
            raise Http404

        return super(EventManageUsers, self).dispatch(request, *args, **kwargs)

    def get_initial(self):
        initial = super(EventManageUsers, self).get_initial()
        if self.request.GET.get('state') is not None and self.request.GET.get('state') == "registrants":
            initial['state'] = 'registered'
        else:
            initial['state'] = 'participant'
        return initial

    def get_context_data(self, **kwargs):

        # DEFAULT OPTIONS FOR LISTING
        state = 'users'
        order_by = 'username'

        # If an option is provided
        if self.request.GET.get('state') is not None:
            if self.request.GET.get('state') in ['users', 'participants', 'registrants']:
                state = self.request.GET.get('state')
        # If an option is provided
        if self.request.GET.get('order_by') is not None:
            if self.request.GET.get('order_by') in ['username', 'last_name', 'surname', 'year']:
                order_by = self.request.GET.get('order_by')

        initial_list_users_form = {
            'state': state,
            'order_by': order_by,
        }

        # Création des forms
        list_users_form = EventListUsersForm(
            initial=initial_list_users_form)

        context = super(EventManageUsers,
                        self).get_context_data(**kwargs)
        context['pk'] = self.se.pk
        context['done'] = self.se.done
        context['price'] = self.se.price
        context['state'] = state
        context['order_by'] = order_by

        context['list_users_form'] = list_users_form
        context['list_weights'] = sorted(self.get_list_weights(
            state), key=lambda item: getattr(item[0], order_by))
        return context

    def form_valid(self, form):
        user = form.cleaned_data['user']
        weight = form.cleaned_data['weight']
        # True pour un participant
        isParticipant = form.cleaned_data['state'] == 'participant'

        self.se.add_weight(user, weight, isParticipant)
        return super(EventManageUsers, self).form_valid(form)

    def get_success_url(self):
        return reverse('url_event_manage_users',
                       kwargs={'group_name': self.group.name, 'pk': self.se.pk})


class EventRemoveUser(GroupPermissionMixin, View):
    perm_codename = None

    def get(self, request, *args, **kwargs):
        se = Event.objects.get(pk=kwargs['pk'])

        # Permission
        if se.done is True:
            raise PermissionDenied
        if not (request.user == se.manager or request.user.has_perm('finances.change_event')):
            raise PermissionDenied

        try:
            state = request.GET['state']
            order_by = request.GET['order_by']
            user_pk = kwargs['user_pk']

            if state == "users":
                if user_pk == 'ALL':
                    for u in se.users.all():
                        se.remove_user(u)
                else:
                    se.remove_user(User.objects.get(pk=user_pk))

            elif state == "participants":
                if user_pk == 'ALL':
                    for u in se.users.all():
                        se.change_weight(u, 0, True)
                else:
                    se.change_weight(User.objects.get(pk=user_pk), 0, True)

            elif state == "registrants":
                if user_pk == 'ALL':
                    for u in se.users.all():
                        se.change_weight(u, 0, False)
                else:
                    se.change_weight(User.objects.get(pk=user_pk), 0, False)

            else:
                raise Http404
        except ObjectDoesNotExist:
            raise Http404

        return redirect(reverse(
            'url_event_manage_users',
            kwargs={'group_name': self.group.name, 'pk': se.pk}
        ) + "?state=" + state + "&order_by=" + order_by + "#table_users")


class EventDownloadXlsx(GroupPermissionMixin, FormView, GroupLateralMenuMixin):
    """
    Download Excel.
    """
    form_class = EventDownloadXlsxForm
    perm_codename = None  # Checked in dispatch()

    def dispatch(self, request, *args, **kwargs):
        try:
            self.group = Group.objects.get(name=kwargs['group_name'])
            self.se = Event.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404

        # Permission
        if not (request.user == self.se.manager or request.user.has_perm('finances.change_event')):
            raise PermissionDenied

        return super(EventDownloadXlsx, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):

        wb = Workbook()
        # grab the active worksheet
        ws = wb.active
        ws.title = "Test"
        ws.append(['Username', 'Pondération',
                   'Infos (Non utilisées) ->', 'Nom Prénom', 'Bucque'])

        if form.cleaned_data['state'] == 'year':

            if form.cleaned_data['years']:
                # Contains the years selected
                list_year_result = form.cleaned_data['years']

                users = User.objects.filter(year__in=list_year_result, is_active=True).exclude(
                    groups=Group.objects.get(pk=1)).order_by('username')
                for u in users:
                    ws.append([u.username, '', '', u.last_name +
                               ' ' + u.first_name, u.surname])

            else:
                raise Http404

        elif form.cleaned_data['state'] == 'participants':
            list_participants_weight = self.se.list_participants_weight()
            for e in list_participants_weight:
                u = e[0]
                ws.append([u.username, e[1], u.last_name +
                           ' ' + u.first_name, u.surname])

        elif form.cleaned_data['state'] == 'registrants':
            list_registrants_weight = self.se.list_registrants_weight()
            for e in list_registrants_weight:
                u = e[0]
                ws.append([u.username, e[1], u.last_name +
                           ' ' + u.first_name, u.surname])
        else:
            raise Http404

        # Return the file
        response = HttpResponse(save_virtual_workbook(
            wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = (
            'attachment; filename="' + self.se.__str__() + '.xlsx"')
        return response


class EventUploadXlsx(GroupPermissionMixin, FormView, GroupLateralMenuMixin):
    """
    Upload Excel.
    """
    form_class = EventUploadXlsxForm
    perm_codename = None  # Checked in dispatch()

    def dispatch(self, request, *args, **kwargs):
        try:
            self.group = Group.objects.get(name=kwargs['group_name'])
            self.se = Event.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404

        # Permission
        if self.se.done is True:
            raise PermissionDenied
        if not (request.user == self.se.manager or request.user.has_perm('finances.change_event')):
            raise PermissionDenied

        return super(EventUploadXlsx, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form, *args, **kwargs):
        try:
            wb = load_workbook(self.request.FILES['list_user'], read_only=True)
            sheet = wb.active
            rows = sheet.rows
            next(rows)  # Skip the first row
            data = []
        except:
            raise PermissionDenied

        if form.cleaned_data['state'] == 'participants':
            isParticipant = True
        else:
            isParticipant = False

        errors = []
        nb_empty_rows = 0
        min_col = sheet.min_column - 1
        i = 1

        # Enregistrement des pondérations
        for row in rows:
            i += 1
            try:
                if row[min_col].value and row[min_col + 1].value:
                    username = row[min_col].value.strip()  # Should be a str
                    if User.objects.filter(username=username).count() > 0:
                        user = User.objects.get(username=username)
                        try:
                            # Should be an int. Else, raise an error
                            pond = int(row[min_col+1].value)
                            if pond > 0:
                                self.se.change_weight(
                                    user, pond, isParticipant)
                        except:
                            errors.append(
                                "Erreur avec " + username + " (ligne n*" + str(i) + "). A priori pas ajouté.")
                    else:
                        errors.append("L'utilisateur " + username +
                                      " n'existe pas. (ligne n*" + str(i) + ").")
                else:
                    nb_empty_rows += 1
            except:
                errors.append("Erreur avec la ligne n*" +
                              str(i) + ". Pas ajouté.")

        errors_count = len(errors)
        error_message = ""
        if errors_count >= 1:
            error_message = str(errors_count) + \
                " erreur(s) pendant l'ajout : \n - "
            if sheet.max_row - nb_empty_rows - 1 == errors_count:
                error_message += "Aucune donnée ne peut être importée (Vérifiez le format et la syntaxe du contenu du fichier)"
            else:
                error_message += "\n - ".join(errors)
            messages.warning(self.request, error_message)
            if i - nb_empty_rows - (1 + errors_count) > 0:
                messages.success(self.request, "Les " + str(i - nb_empty_rows -
                                                            (1 + errors_count)) + " autres utilisateurs ont bien été ajoutés.")
        else:
            messages.success(self.request, "Les " + str(i-1) +
                             " utilisateurs ont bien été ajoutés.")

        return super(EventUploadXlsx, self).form_valid(form)

    def get_success_url(self):
        return reverse('url_event_manage_users',
                       kwargs={'group_name': self.group.name, 'pk': self.se.pk})
