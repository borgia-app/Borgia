import datetime
import json
import random
import re
import string

import openpyxl
from django.contrib import messages
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.contrib.auth.models import Group, Permission
from django.contrib.messages.views import SuccessMessageMixin
from django.core.exceptions import ObjectDoesNotExist, PermissionDenied
from django.db.models import Q
from django.http import Http404, HttpResponseBadRequest
from django.shortcuts import HttpResponse, redirect, render
from django.urls import reverse
from django.utils.encoding import force_text
from django.views.generic.base import View
from django.views.generic.edit import FormView

from borgia.utils import (GroupPermissionMixin, LateralMenuMixin,
                          get_members_group, group_name_display,
                          human_unused_permissions, permission_to_manage_group)
from configurations.utils import configurations_safe_get
from events.models import Event
from users.forms import (GroupUpdateForm, SelfUserUpdateForm,
                         UserCreationCustomForm, UserDownloadXlsxForm,
                         UserSearchForm, UserUpdateForm, UserUploadXlsxForm)
from users.mixins import UserMixin
from users.models import User


class UserListView(PermissionRequiredMixin, FormView, LateralMenuMixin):
    """
    List User instances.

    """
    permission_required = 'users.view_user'
    form_class = UserSearchForm
    template_name = 'users/user_list.html'
    lm_active = 'lm_user_list'

    search = None
    year = None
    state = None
    headers = {'username': 'asc',
               'last_name': 'asc',
               'surname': 'asc',
               'family': 'asc',
               'campus': 'asc',
               'year': 'asc',
               'balance': 'asc'}
    sort = None

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Header List
        context['list_header'] = [["username", "Username"], ["last_name", "Nom Prénom"], ["surname", "Bucque"], [
            "family", "Fam's"], ["campus", "Tabagn's"], ["year", "Prom's"], ["balance", "Solde"]]

        try:
            self.sort = self.request.GET['sort']
        except KeyError:
            pass

        if self.sort is not None:
            context['sort'] = self.sort
            if self.headers[self.sort] == "des":
                context['reverse'] = True
                context['user_list'] = self.form_query(
                    User.objects.all().exclude(groups=1).order_by(self.sort).reverse())
                self.headers[self.sort] = "asc"
            else:
                context['user_list'] = self.form_query(
                    User.objects.all().exclude(groups=1).order_by(self.sort))
                self.headers[self.sort] = "des"
        else:
            context['user_list'] = self.form_query(
                User.objects.all().exclude(groups=1))

        return context

    def form_query(self, query):
        if self.search:
            query = query.filter(
                Q(last_name__icontains=self.search)
                | Q(first_name__icontains=self.search)
                | Q(surname__icontains=self.search)
                | Q(username__icontains=self.search)
            )

        if self.year and self.year != 'all':
            query = query.filter(
                year=self.year)

        if self.state and self.state != 'all':
            if self.state == 'negative_balance':
                query = query.filter(balance__lt=0.0, is_active=True)
            elif self.state == 'threshold':
                threshold = configurations_safe_get(
                    'BALANCE_THRESHOLD_PURCHASE').get_value()
                query = query.filter(balance__lt=threshold, is_active=True)
            elif self.state == 'unactive':
                query = query.filter(is_active=False)
        else:
            query = query.filter(is_active=True)

        return query

    def form_valid(self, form):
        if form.cleaned_data['search']:
            self.search = form.cleaned_data['search']

        if form.cleaned_data['state']:
            self.state = form.cleaned_data['state']

        if form.cleaned_data['year']:
            self.year = form.cleaned_data['year']

        return self.get(self.request, self.args, self.kwargs)

    def get_initial(self):
        initial = super().get_initial()
        initial['search'] = self.search
        return initial


class UserCreateView(PermissionRequiredMixin, SuccessMessageMixin, FormView, LateralMenuMixin):
    """
    Create a new user and redirect to the workboard of the group.

    """
    permission_required = 'users.add_user'
    form_class = UserCreationCustomForm
    template_name = 'users/user_create.html'
    lm_active = 'lm_user_create'
    success_url = None

    def form_valid(self, form):
        user = User.objects.create(username=form.cleaned_data['username'],
                                   first_name=form.cleaned_data['first_name'],
                                   last_name=form.cleaned_data['last_name'],
                                   email=form.cleaned_data['email'],
                                   surname=form.cleaned_data['surname'],
                                   family=form.cleaned_data['family'],
                                   campus=form.cleaned_data['campus'],
                                   year=form.cleaned_data['year'])
        user.set_password(form.cleaned_data['password'])
        user.save()

        is_external_member = form.cleaned_data['is_external_member']
        user.groups.add(get_members_group(is_external_member))

        user.save()

        # User object is assigned to self.object (so we can access to it in get_success_url)
        self.object = user

        return super().form_valid(form)

    def get_initial(self):
        initial = super().get_initial()
        initial['campus'] = 'Me'
        initial['year'] = datetime.datetime.now().year - 1
        return initial

    def get_success_message(self, cleaned_data):
        return "L'utilisateur a bien été crée'"

    def get_success_url(self):
        """
        If can retrieve user: go to the user.
        If not, go to the workboard of the group.
        """
        if self.request.user.has_perm('users.view_user'):
            return reverse('url_user_retrieve', kwargs={'user_pk': self.object.pk})
        else:
            return reverse('url_managers_workboard')


class UserRetrieveView(UserMixin, View, LateralMenuMixin):
    """
    Retrieve a User instance.

    """
    permission_required = 'users.view_user'
    template_name = 'users/user_retrieve.html'

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        return render(request, self.template_name, context=context)


class UserUpdateView(UserMixin, SuccessMessageMixin, FormView, LateralMenuMixin):
    """
    Update an user and redirect to the workboard of the group.

    """
    permission_required = 'users.change_user'
    form_class = UserUpdateForm
    template_name = 'users/user_update.html'
    model = User
    modified = False

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.user
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        user = self.user
        for k in UserUpdateForm(user=user).fields.keys():
            initial[k] = getattr(user, k)
        return initial

    def form_valid(self, form):
        user = self.user
        for k in form.fields.keys():
            if form.cleaned_data[k] != getattr(user, k):
                self.modified = True
                setattr(user, k, form.cleaned_data[k])
        user.save()

        return super().form_valid(form)

    def get_success_message(self, cleaned_data):
        return "Les informations ont bien été mises à jour"

    def get_success_url(self):
        return reverse('url_user_retrieve',
                       kwargs={'user_pk': self.user.pk})


class UserDeactivateView(UserMixin, View, LateralMenuMixin):
    """
    Deactivate a user and redirect to the workboard of the group.

    """
    permission_required = 'users.delete_user'
    template_name = 'users/user_deactivate.html'
    success_message = "Le compte de %(user)s a bien été "
    error_event_message = "Veuillez attribuer la gestion des évènements suivants à un autre utilisateur avant de désactiver le compte:"

    def has_permission(self):
        has_perm = super().has_permission()
        if has_perm:
            return True
        else:
            return self.user == self.request.user

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        return render(request, self.template_name, context=context)

    def post(self, request, *args, **kwargs):
        deactivated = False
        if self.user.is_active is True:
            events = self.user.manager.filter(done=False)
            if events.count() > 0:
                for event in events:
                    self.error_event_message += "\n - " + event.description
                messages.warning(request, self.error_event_message)
            else:
                deactivated = True
                self.user.is_active = False
                # si c'est un gadz. Special members can't be added to other groups
                if Group.objects.get(pk=5) in self.user.groups.all():
                    self.user.groups.clear()
                    self.user.groups.add(Group.objects.get(pk=5))
                self.user.save()            
        else:
            self.user.is_active = True
        self.user.save()

        if self.user.is_active:
            self.success_message += 'activé'
        else:
            self.success_message += 'désactivé'

        messages.success(request, self.success_message % dict(
            user=self.user,
        ))

        if request.user == self.user and deactivated:
             self.success_url = reverse(
                'url_logout')
        else:           
            self.success_url = reverse(
                'url_user_retrieve',
                kwargs={'user_pk': self.user.pk})

        return redirect(force_text(self.success_url))


class UserSelfDeactivateView(GroupPermissionMixin, View, LateralMenuMixin):
    """
    Deactivate own account and disconnect.

    :param kwargs['group_name']: name of the group used.
    :param self.perm_codename: codename of the permission checked.
    """
    template_name = 'users/deactivate.html'
    perm_codename = None
    error_event_message = "Veuillez attribuer la gestion des évènements suivants à un autre utilisateur avant de désactiver le compte:"

    def get(self, request, *args, **kwargs):
        try:
            user = User.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404
        context = self.get_context_data(**kwargs)
        context['user'] = user
        return render(request, 'users/deactivate.html', context=context)

    def post(self, request, *args, **kwargs):
        try:
            user = User.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404
        events = Event.objects.filter(manager=user, done=False)
        if user.is_active is True:
            if events.count() > 0:
                for event in events:
                    self.error_event_message += "\n - " + event.description
                messages.warning(request, self.error_event_message)
            else:
                user.is_active = False
                # si c'est un gadz. Special members can't be added to other groups
                if Group.objects.get(pk=5) in user.groups.all():
                    user.groups.clear()
                    user.groups.add(Group.objects.get(pk=5))
                user.save()
        if events.count() > 0:
            self.success_url = reverse(
                'url_event_list',
                kwargs={'group_name': self.group.name})
        else:
            self.success_url = reverse(
                'url_logout')
        return redirect(force_text(self.success_url))


class UserSelfUpdateView(SuccessMessageMixin, FormView, LateralMenuMixin):
    template_name = 'users/user_self_update.html'
    form_class = SelfUserUpdateForm

    def get_initial(self):
        initial = super().get_initial()
        initial['email'] = self.request.user.email
        initial['phone'] = self.request.user.phone
        initial['avatar'] = self.request.user.avatar
        initial['theme'] = self.request.user.theme
        return initial

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        self.request.user.email = form.cleaned_data['email']
        self.request.user.phone = form.cleaned_data['phone']
        self.request.user.theme = form.cleaned_data['theme']
        if form.cleaned_data['avatar'] is not False:
            setattr(self.request.user, 'avatar', form.cleaned_data['avatar'])
        else:
            if self.request.user.avatar:
                self.request.user.avatar.delete(True)
        self.request.user.save()
        return super().form_valid(form)

    def get_success_message(self, cleaned_data):
        return "Vos infos ont bien été mises à jour"


class GroupUpdateView(PermissionRequiredMixin, SuccessMessageMixin, FormView,
                      LateralMenuMixin):
    template_name = 'users/group_update.html'
    success_url = None
    form_class = GroupUpdateForm
    group_updated = None
    lm_active = None

    def get_permission_required(self):
        """
        Override the permission_required attribute.
        Must return an iterable.
        """
        #TODO : Get group
        # TODO : move dispatch into get_initial and context

        try:
            self.group = Group.objects.get(pk=self.kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404
        perms = (permission_to_manage_group(self.group_updated)[1],)
        self.lm_active = 'lm_group_manage_' + self.group.name
        return perms

    def get_form_kwargs(self):
        """
        Add possible members and permissions to kwargs of the form.

        Possible members are all members, except externals members and unactive users.
        Possible permissions are all permissions.
        :note:: For the special case of a shop management, two groups exist:
        group of chiefs and group of associates. If the group of associates is
        managed, possible permissions are only permissions of the chiefs group.
        """
        kwargs = super().get_form_kwargs()

        if self.group.name.startswith('associates-') is True:
            chiefs_group_name = self.group.name.replace(
                'associates', 'chiefs')
            kwargs['possible_permissions'] = Permission.objects.filter(
                pk__in=[p.pk for p in Group.objects.get(
                    name=chiefs_group_name).permissions.all().exclude(
                    pk=permission_to_manage_group(self.group)[0].pk).exclude(
                    pk__in=human_unused_permissions())]
            )

        else:
            kwargs['possible_permissions'] = Permission.objects.all().exclude(
                pk__in=human_unused_permissions()
            )

        kwargs['possible_members'] = User.objects.filter(is_active=True).exclude(
            groups=Group.objects.get(name='externals'))
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        initial['members'] = self.group.user_set.all()
        initial['permissions'] = self.group.permissions.all()
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['group'] = self.group
        return context

    def form_valid(self, form):
        """
        Update permissions and members of the group updated.
        """
        old_members = User.objects.filter(groups=self.group)
        new_members = form.cleaned_data['members']
        old_permissions = self.group.permissions.all()
        new_permissions = form.cleaned_data['permissions']

        # Modification des membres
        for m in old_members:
            if m not in new_members:
                m.groups.remove(self.group)
                m.save()
        for m in new_members:
            if m not in old_members:
                m.groups.add(self.group)
                m.save()

        # Modification des permissions
        for p in old_permissions:
            if p not in new_permissions:
                self.group.permissions.remove(p)
        for p in new_permissions:
            if p not in old_permissions:
                self.group.permissions.add(p)
        self.group.save()

        return super().form_valid(form)

    def get_success_message(self, cleaned_data):
        return "Le groupe a bien été mis à jour"

    def get_success_url(self):
        return reverse('url_managers_workboard')


class UserUploadXlsxView(GroupPermissionMixin, FormView, LateralMenuMixin):
    """
    Download/Upload Excel for adding users.
    """
    template_name = 'users/bulk_edit.html'
    form_class = UserUploadXlsxForm
    second_form_class = UserDownloadXlsxForm
    perm_codename = 'add_user'
    lm_active = 'lm_user_create'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'upload_form' not in context:
            context['upload_form'] = self.form_class()
        if 'download_form' not in context:
            context['download_form'] = self.second_form_class()
        return context

    def form_valid(self, form, *args, **kwargs):

        try:
            wb = openpyxl.load_workbook(self.request.FILES['list_user'], read_only=True)
            sheet = wb.active
            rows = sheet.rows
        except:
            raise PermissionDenied

        errors = []
        nb_empty_rows = 0
        min_col = sheet.min_column
        max_col = sheet.max_column
        min_row = sheet.min_row
        i = min_row

        columns = form.cleaned_data['xlsx_columns']
        # Setting column numbers
        for col in range(min_col, max_col+1):
            if sheet.cell(None, sheet.min_row, col).value == 'username':
                col_username = col - min_row
            elif sheet.cell(None, sheet.min_row, col).value == 'first_name':
                col_first_name = col - min_row
            elif sheet.cell(None, sheet.min_row, col).value == 'last_name':
                col_last_name = col - min_row
            elif sheet.cell(None, sheet.min_row, col).value == 'email':
                col_email = col - min_row
            elif sheet.cell(None, sheet.min_row, col).value == 'surname':
                col_surname = col - min_row
            elif sheet.cell(None, sheet.min_row, col).value == 'family':
                col_family = col - min_row
            elif sheet.cell(None, sheet.min_row, col).value == 'campus':
                col_campus = col - min_row
            elif sheet.cell(None, sheet.min_row, col).value == 'year':
                col_year = col - min_row
            elif sheet.cell(None, sheet.min_row, col).value == 'balance':
                col_balance = col - min_row

        for _ in range(min_row):
            next(rows)

        skipped_rows = []
        # Saving users
        for row in rows:
            i += 1
            try:
                user_dict = {}
                errors_on_required_columns = []
                skipped_row = False
                try:
                    if row[col_username].value:
                        user_dict['username'] = row[col_username].value.strip()
                    else:
                        skipped_row = True
                        nb_empty_rows += 1
                except:
                    skipped_row = True
                    errors_on_required_columns.append('username')

                if 'first_name' in columns:
                    try:
                        if row[col_first_name].value:
                            user_dict['first_name'] = row[col_first_name].value.strip()
                    except:
                        errors_on_required_columns.append('first_name')

                if 'last_name' in columns:
                    try:
                        if row[col_last_name].value:
                            user_dict['last_name'] = row[col_last_name].value.strip()
                    except:
                        errors_on_required_columns.append('last_name')

                if 'email' in columns:
                    try:
                        if row[col_email].value:
                            user_dict['email'] = row[col_email].value.strip()
                    except:
                        errors_on_required_columns.append('email')

                if 'surname' in columns:
                    try:
                        if row[col_surname].value:
                            user_dict['surname'] = row[col_surname].value.strip()
                    except:
                        errors_on_required_columns.append('surname')

                if 'family' in columns:
                    try:
                        if row[col_family].value:
                            user_dict['family'] = str(row[col_family].value).strip()
                    except:
                        errors_on_required_columns.append('family')

                if 'campus' in columns:
                    try:
                        if row[col_campus].value:
                            user_dict['campus'] = row[col_campus].value.strip()
                    except:
                        errors_on_required_columns.append('campus')

                if 'year' in columns:
                    try:
                        if row[col_year].value:
                            user_dict['year'] = int(row[col_year].value)
                    except:
                        errors_on_required_columns.append('year')

                if 'balance' in columns:
                    try:
                        if row[col_balance].value:
                            user_dict['balance'] = str(row[col_balance].value).strip()
                    except:
                        errors_on_required_columns.append('balance')

                if not skipped_row:
                    username = user_dict['username']
                    if User.objects.filter(username=username).count() > 0:
                        User.objects.filter(username=username).update(**user_dict)
                    else:
                        user = User.objects.create(**user_dict)
                        user.set_password(''.join(random.SystemRandom().choice(string.ascii_uppercase + string.digits)
                                                  for _ in range(20)))
                    user = User.objects.get(username=username)
                    user.save()

                    user.groups.add(Group.objects.get(pk=5))
                else:
                    skipped_rows.append(str(i))
            except:
                if str(i) in skipped_rows:
                    errors.append("La ligne n*" + str(i) +
                                  "n'a pas été traitée car le username est manquant")  # TODO: Make this error appear
                elif len(errors_on_required_columns) > 0:
                    errors.append("Les colonnes " + ", ".join(errors_on_required_columns) +
                                  " sont requis et comportent des erreurs (ligne n*" + str(i) + ")")
                else:
                    errors.append("Une erreur empêche le traitement du fichier Excel (ligne n*" + str(i) + ")")

        error_message = ""
        error_message += "\n - ".join(errors)
        messages.success(self.request, str(i - len(skipped_rows) - min_row - len(errors)) +
                         " utilisateurs ont été crées/mis à jour")
        messages.warning(self.request, error_message)

        return super().form_valid(form)

    def get_success_url(self):
        try:
            if Permission.objects.get(codename='list_user') in self.group.permissions.all():
                return reverse('url_user_list',
                               kwargs={'group_name': self.group.name})
            else:
                return reverse('url_group_workboard',
                               kwargs={'group_name': self.group.name})
        except ObjectDoesNotExist:
            return reverse('url_workboard',
                           kwargs={'group_name': self.group.name})


class UserAddByListXlsxDownload(GroupPermissionMixin, View, LateralMenuMixin):
    """
    Download Excel for adding users.
    """
    perm_codename = 'add_user'
    lm_active = 'lm_user_create'

    def get(self, request, *args, **kwargs):
        columns = request.GET.getlist('xlsx_columns')

        if len(columns) == 0 or columns is None:
            columns = ['first_name', 'last_name', 'email', 'surname',
                       'family', 'campus', 'year', 'balance']

        columns.insert(0, 'username')

        wb = openpyxl.Workbook()
        # grab the active worksheet
        ws = wb.active
        ws.title = "users"
        ws.append(columns)

        # Return the file
        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb),
                                content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="UsersList.xlsx"'

        users = User.objects.all().values_list(*columns)
        for user in users:
            ws.append(user)

        wb.save(response)

        return response


def username_from_username_part(request):
    data = []

    try:
        key = request.GET.get('keywords')

        regex = r"^" + re.escape(key) + r"(\W|$)"

        # Fam'ss en entier
        # where_search = User.objects.filter(family=key).exclude(groups=1).order_by('-year')
        where_search = User.objects.exclude(groups=1).filter(
            family__regex=regex, is_active=True).order_by('-year')

        if len(key) > 2:
            if key.isalpha():
                # Nom de famille, début ou entier à partir de 3 caractères
                where_search = where_search | User.objects.filter(
                    last_name__istartswith=key, is_active=True)
                # Prénom, début ou entier à partir de 3 caractères
                where_search = where_search | User.objects.filter(
                    first_name__istartswith=key, is_active=True)
                # Buque, début ou entier à partir de 3 caractères
                where_search = where_search | User.objects.filter(
                    surname__istartswith=key, is_active=True)

                # Suppression des doublons
                where_search = where_search.distinct()

        for e in where_search:
            data.append(e.username)

    except KeyError:
        pass

    return HttpResponse(json.dumps(data))


def balance_from_username(request):
    # Check permissions

    # User is authentified, if not the he can't access the view
    operator = request.user

    # try:
    #     shop_name = request.GET.get('shop_name')
    #     shop = Shop.objects.get(name = shop_name)
    #     module = OperatorSaleModule.objects.get(shop = shop)
    # except KeyError:
    #         raise Http404
    # except ObjectDoesNotExist:
    #     raise Http404

    # If deactivate
    # if module.state is False:
    #     raise Http404

    if operator.has_perm('modules.use_operatorsalemodule'):
        try:
            username = request.GET['username']
            data = str(User.objects.get(username=username).balance)
            return HttpResponse(json.dumps(data))
        except KeyError:
            return HttpResponseBadRequest()
        except ObjectDoesNotExist:
            return HttpResponseBadRequest()

    # If user don't have the permission
    else:
        raise PermissionDenied
