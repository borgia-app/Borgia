import datetime
import decimal

from django.contrib import messages
from django.contrib.auth.mixins import (LoginRequiredMixin,
                                        PermissionRequiredMixin)
from django.contrib.auth.models import Group, Permission
from django.core.exceptions import ObjectDoesNotExist, PermissionDenied
from django.http import Http404
from django.shortcuts import HttpResponse, redirect
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from borgia.utils import get_members_group
from borgia.views import BorgiaFormView, BorgiaView
from events.forms import (EventAddWeightForm, EventCreateForm, EventDeleteForm,
                          EventDownloadXlsxForm, EventFinishForm,
                          EventListForm, EventListUsersForm,
                          EventSelfRegistrationForm, EventUpdateForm,
                          EventUploadXlsxForm)
from events.mixins import EventMixin
from events.models import Event
from users.models import User


class EventList(LoginRequiredMixin, PermissionRequiredMixin, BorgiaFormView):
    permission_required = 'events.view_event'
    menu_type = 'members'
    template_name = 'events/event_list.html'
    lm_active = 'lm_event_list'
    form_class = EventListForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        events = Event.objects.filter(
            date__gte=datetime.date.today().replace(day=1), done=False).order_by('-date')

        for event in events:  # Duplicate
            # Si fini, on recupere la participation, sinon la preinscription
            event.weight_of_user = event.get_weight_of_user(
                self.request.user, event.done)
            event.number_registrants = event.get_number_registrants()
            event.number_participants = event.get_number_participants()
            event.total_weights_registrants = event.get_total_weights_registrants()
            event.total_weights_participants = event.get_total_weights_participants()
            try:
                event.has_perm_manage = (self.request.user == event.manager or
                                         self.request.user.has_perm('events.change_event'))
            except ObjectDoesNotExist:
                event.has_perm_manage = False
        context['events'] = events
        # Permission SelfRegistration
        if self.request.user.has_perm('events.self_register_event'):
            context['has_perm_self_register_event'] = True

        return context

    def form_valid(self, form, **kwargs):
        date_begin = form.cleaned_data['date_begin']
        date_end = form.cleaned_data['date_end']
        order_by = form.cleaned_data['order_by']
        done = form.cleaned_data['done']

        if date_begin is None:
            date_begin = datetime.date.min
        if date_end is None:
            date_end = datetime.date.max

        if done != "both":
            if done == "done":
                done = True
            else:
                done = False

            events = Event.objects.filter(
                date__range=[date_begin, date_end], done=done)
        else:
            events = Event.objects.filter(
                date__range=[date_begin, date_end])

        context = self.get_context_data(**kwargs)
        if order_by != '-date':
            events = events.order_by(order_by).order_by('-date')
        else:
            events = events.order_by('-date')

        for event in events:  # Duplicate
            # Si fini, on recupere la participation, sinon la preinscription
            event.weight_of_user = event.get_weight_of_user(
                self.request.user, event.done)
            event.number_registrants = event.get_number_registrants()
            event.number_participants = event.get_number_participants()
            event.total_weights_registrants = event.get_total_weights_registrants()
            event.total_weights_participants = event.get_total_weights_participants()

        context['events'] = events

        return self.render_to_response(context)


class EventCreate(LoginRequiredMixin, PermissionRequiredMixin, BorgiaFormView):
    permission_required = 'events.add_event'
    menu_type = 'members'
    success_message = "'%(description)s' a bien été créé."
    template_name = 'events/event_create.html'
    form_class = EventCreateForm
    lm_active = 'lm_event_create'

    def __init__(self):
        self.event = None

    def form_valid(self, form):
        event = Event.objects.create(
            description=form.cleaned_data['description'],
            date=form.cleaned_data['date'],
            allow_self_registeration=form.cleaned_data['allow_self_registeration'],
            manager=self.request.user)
        if form.cleaned_data['price']:
            event.price = decimal.Decimal(form.cleaned_data['price'])
        if form.cleaned_data['bills']:
            event.bills = form.cleaned_data['bills']
        if form.cleaned_data['date_end_registration']:
            date_end_registration = form.cleaned_data['date_end_registration']
            if date_end_registration > datetime.date.today() and date_end_registration <= form.cleaned_data['date']:
                event.date_end_registration = form.cleaned_data['date_end_registration']

        event.save()
        self.event = event

        return super().form_valid(form)

    def get_success_message(self, cleaned_data):
        return self.success_message % dict(
            description=self.event.description,
        )

    def get_success_url(self):
        return reverse('url_event_update',
                       kwargs={'pk': self.event.pk}
                       )


class EventUpdate(EventMixin, BorgiaFormView):
    """
    Update the Shared Event
    """
    permission_required = 'events.change_event'
    menu_type = 'members'
    success_message = "'%(description)s' a bien été mis à jour."
    template_name = 'events/event_update.html'
    form_class = EventUpdateForm
    allow_manager = True

    manager_changed = False

    def get_initial(self):
        initial = super().get_initial()
        initial['price'] = self.event.price
        initial['bills'] = self.event.bills
        initial['manager'] = self.event.manager.username
        initial['allow_self_registeration'] = self.event.allow_self_registeration
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Pour les users
        number_participants = self.event.get_number_participants()
        context['number_registrants'] = self.event.get_number_registrants()
        context['number_participants'] = number_participants
        context['total_weights_registrants'] = self.event.get_total_weights_registrants()
        context['total_weights_participants'] = self.event.get_total_weights_participants()
        context['no_participant'] = number_participants == 0

        # Création des forms excel
        context['upload_xlsx_form'] = EventUploadXlsxForm()
        context['download_xlsx_form'] = EventDownloadXlsxForm()

        # Permission FinishEvent
        if self.request.user.has_perm('events.proceed_payment_event'):
            context['has_perm_proceed_payment'] = True

        return context

    def form_valid(self, form):
        if form.cleaned_data['price']:
            self.event.price = form.cleaned_data['price']
        if form.cleaned_data['bills']:
            self.event.bills = form.cleaned_data['bills']
        if form.cleaned_data['manager']:
            form_manager = User.objects.get(
                username=form.cleaned_data['manager'])
            manage_permission = False
            if self.event.manager != form_manager:
                for group in form_manager.groups.all():
                    if Permission.objects.get(codename='add_event') in group.permissions.all():
                        manage_permission = True
                        break
                if manage_permission:
                    self.event.manager = form_manager
                    self.manager_changed = True
                else:
                    messages.warning(self.request,
                                     "%(user)s ne dispose pas de droits suffisants pour gérer l'évènement" % dict(
                                         user=form_manager))
        self.event.allow_self_registeration = form.cleaned_data['allow_self_registeration']
        self.event.save()

        return super().form_valid(form)

    def get_success_message(self, cleaned_data):
        if self.manager_changed:
            return "%(user)s gère désormais l'évènement" % dict(
                user=self.event.manager,
            )
        else:
            return self.success_message % dict(
                description=self.event.description,
            )

    def get_success_url(self):
        if self.manager_changed and not self.request.user.has_perm('events.change_event'):
            return reverse('url_members_workboard')
        else:
            return reverse('url_event_update',
                           kwargs={'pk': self.event.pk}
                           )


class EventFinish(EventMixin, BorgiaFormView):
    """
    Finish a event and redirect to the list of events.
    This command is used when you want to keep the event in the database, but
    you don't want to pay in Borgia (for instance paid with real money).

    :param kwargs['group_name']: name of the group used.
    :param kwargs['pk']: pk of the event
    :param self.perm_codename: codename of the permission checked.
    """
    permission_required = 'events.proceed_payment_event'
    menu_type = 'managers'
    success_message = "'%(description)s' a bien été terminé."
    template_name = 'events/event_finish.html'
    form_class = EventFinishForm
    need_ongoing_event = True


    def __init__(self):
        super().__init__()
        self.total_weights_participants = None
        self.ponderation_price = None

    def has_permission(self):
        """
        Check if event exists, then permission.
        Then check potentially on-going / manager attributes.
        Then, check if there at least one participant.
        """
        if not super().has_permission():
            return False
        else:
            self.total_weights_participants = self.event.get_total_weights_participants()
            if self.total_weights_participants == 0:
                return False

            try:
                self.ponderation_price = round(self.event.price / self.total_weights_participants, 2)
            except TypeError:
                self.ponderation_price = 0
            return True

    def get_initial(self):
        """
        Populate the form with the current price.
        """
        initial = super().get_initial()
        initial['total_price'] = self.event.price
        initial['ponderation_price'] = self.ponderation_price
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_price'] = self.event.price
        context['total_weights_participants'] = self.total_weights_participants
        context['ponderation_price'] = self.ponderation_price
        return context

    def form_valid(self, form):
        type_payment = form.cleaned_data['type_payment']
        if type_payment == 'pay_by_total':
            self.event.pay_by_total(self.request.user, User.objects.get(
                pk=1), form.cleaned_data['total_price'])

        if type_payment == 'pay_by_ponderation':
            self.event.pay_by_ponderation(self.request.user, User.objects.get(
                pk=1), form.cleaned_data['ponderation_price'])

        if type_payment == 'no_payment':
            self.event.end_without_payment(form.cleaned_data['remark'])

        return super().form_valid(form)

    def get_success_message(self, cleaned_data):
        return self.success_message % dict(
            description=self.event.description,
        )

    def get_success_url(self):
        return reverse('url_event_list')


class EventDelete(EventMixin, BorgiaFormView):
    """
    Delete a event and redirect to the list of events.

    :param kwargs['pk']: pk of the event
    """
    permission_required = 'events.delete_event'
    menu_type = 'members'
    success_message = "L'évènement a bien été supprimé."
    template_name = 'events/event_delete.html'
    form_class = EventDeleteForm
    allow_manager = True
    need_ongoing_event = True

    def form_valid(self, form):
        self.event.delete()
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('url_event_list')


class EventSelfRegistration(EventMixin, BorgiaFormView):
    """
    Allow a user to register himself

    :param kwargs['group_name']: name of the group used.
    :param kwargs['pk']: pk of the event
    :param self.perm_codename: codename of the permission checked.
    """
    permission_required = 'events.self_register_event'
    menu_type = 'members'
    template_name = 'events/event_self_registration.html'
    form_class = EventSelfRegistrationForm
    need_ongoing_event = True

    def __init__(self):
        super().__init__()
        self.new_weight = None

    def has_permission(self):
        """
        Check permission, and if event exists, and if it is not done yet.
        Then, it checks if the event allow_self_registeration, and if it is not an old event.
        """
        has_perms = super().has_permission()
        if not has_perms:
            return False
        else:
            if not self.event.allow_self_registeration:
                return False
            else:
                if self.event.date_end_registration:
                    if datetime.date.today() > self.event.date_end_registration:
                        return False
                return True

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['registeration_of_user'] = self.event.get_weight_of_user(
            self.request.user, False)  # Duplicate
        return context

    def get_initial(self):
        initial = super().get_initial()
        initial['weight'] = self.event.get_weight_of_user(
            self.request.user, False)  # Duplicate
        return initial

    def form_valid(self, form):
        self.new_weight = int(form.cleaned_data['weight'])
        self.event.change_weight(self.request.user, self.new_weight, False)
        return super().form_valid(form)

    def get_success_message(self, cleaned_data):
        if self.new_weight > 0:
            return "Vous avez bien été inscrit"
        else:
            return "Vous avez bien été désinscrit"

    def get_success_url(self):
        return reverse('url_event_self_registration',
                       kwargs={'pk': self.event.pk}
                       )


class EventChangeWeight(EventMixin, BorgiaView):
    permission_required = 'events.change_event'
    menu_type = 'members'
    allow_manager = True
    need_ongoing_event = True

    def get(self, request, *args, **kwargs):
        """
        Change la valeur de la pondération d'un participant user pour un événement
        Permissions :   Si événements terminé -> denied,
                        Si pas manager ou pas la perm 'finances.change_event' -> denied
        :param pk: pk de l'événement
        :param user_pk: paramètre GET correspondant au pk de l'user
        :param pond_pk: paramètre GET correspondant à la nouvelle pondération
        :type pk, user_pk, pond_pk: int
        """
        response = 0

        try:
            user = User.objects.get(pk=kwargs['user_pk'])
            pond = int(request.GET['pond'])
            is_participant = int(request.GET['is_participant'])
        except KeyError:
            pass
        except ObjectDoesNotExist:
            pass
        except ValueError:
            pass

        if pond > 0:
            if is_participant in [0, 1]:
                # Changement de la pondération
                self.event.change_weight(user, pond, is_participant)

                # Réponse
                response = 1

        return HttpResponse(response)


class EventManageUsers(EventMixin, BorgiaFormView):
    """
    Manage the users. The get displays the list of concerned users. The post form add weight to one of them.
    """
    permission_required = 'events.change_event'
    menu_type = 'members'
    template_name = 'events/event_manage_users.html'
    form_class = EventAddWeightForm
    allow_manager = True
    need_ongoing_event = True

    def get_list_weights(self, state):
        if state == 'users':
            return self.event.list_users_weight()
        elif state == 'participants':
            return self.event.list_participants_weight()
        elif state == 'registrants':
            return self.event.list_registrants_weight()

    def get_initial(self):
        initial = super().get_initial()
        if self.request.GET.get('state') is not None and self.request.GET.get('state') == "registrants":
            initial['state'] = 'registered'
        else:
            initial['state'] = 'participant'
        return initial

    def get_context_data(self, **kwargs):

        # DEFAULT OPTIONS FOR LISTING
        state = 'users'
        order_by = 'username'

        # If an option is provided
        if self.request.GET.get('state') is not None:
            if self.request.GET.get('state') in ['users', 'participants', 'registrants']:
                state = self.request.GET.get('state')
        # If an option is provided
        if self.request.GET.get('order_by') is not None:
            if self.request.GET.get('order_by') in ['username', 'last_name', 'surname', 'year']:
                order_by = self.request.GET.get('order_by')

        initial_list_users_form = {
            'state': state,
            'order_by': order_by,
        }

        # Création des forms
        list_users_form = EventListUsersForm(
            initial=initial_list_users_form)

        context = super().get_context_data(**kwargs)
        context['pk'] = self.event.pk
        context['done'] = self.event.done
        context['price'] = self.event.price
        context['state'] = state
        context['order_by'] = order_by

        context['list_users_form'] = list_users_form
        context['list_weights'] = sorted(self.get_list_weights(
            state), key=lambda item: getattr(item[0], order_by))
        return context

    def form_valid(self, form):
        user = form.cleaned_data['user']
        weight = form.cleaned_data['weight']
        # True pour un participant
        is_participant = form.cleaned_data['state'] == 'participant'

        self.event.add_weight(user, weight, is_participant)
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('url_event_manage_users',
                       kwargs={'pk': self.event.pk})


class EventRemoveUser(EventMixin, BorgiaView):
    """
    Remove a user
    """
    permission_required = 'events.change_event'
    menu_type = 'members'
    allow_manager = True
    need_ongoing_event = True

    def get(self, request, *args, **kwargs):
        try:
            state = request.GET['state']
            order_by = request.GET['order_by']
            user_pk = kwargs['user_pk']

            if state == "users":
                if user_pk == 'ALL':
                    for u in self.event.users.all():
                        self.event.remove_user(u)
                else:
                    self.event.remove_user(User.objects.get(pk=user_pk))

            elif state == "participants":
                if user_pk == 'ALL':
                    for u in self.event.users.all():
                        self.event.change_weight(u, 0, True)
                else:
                    self.event.change_weight(
                        User.objects.get(pk=user_pk), 0, True)

            elif state == "registrants":
                if user_pk == 'ALL':
                    for u in self.event.users.all():
                        self.event.change_weight(u, 0, False)
                else:
                    self.event.change_weight(
                        User.objects.get(pk=user_pk), 0, False)

            else:
                raise Http404
        except ObjectDoesNotExist:
            raise Http404

        return redirect(reverse(
            'url_event_manage_users',
            kwargs={'pk': self.event.pk}
        ) + "?state=" + state + "&order_by=" + order_by + "#table_users")


class EventDownloadXlsx(EventMixin, LoginRequiredMixin, BorgiaView):
    """
    Download Excel.
    """
    permission_required = 'events.change_event'
    menu_type = 'members'
    allow_manager = True

    def post(self, request, *args, **kwargs):

        wb = Workbook()
        # grab the active worksheet
        ws = wb.active
        ws.title = "event"
        columns = ['Username', 'Pondération',
                   'Infos (Non utilisées) ->', 'Nom Prénom', 'Bucque']
        ws.append(columns)
        for col in ['A','B','C','D','E']:
            ws.column_dimensions[col].width = 30

        state = request.POST.get("state", "")
        years = request.POST.getlist("years", "")

        if state == 'year':

            if years:
                # Contains the years selected
                list_year_result = years
                users = User.objects.filter(year__in=list_year_result, is_active=True).exclude(
                    groups=get_members_group(is_externals=True)).order_by('-year','username')
                for u in users:
                    ws.append([u.username, '', '', u.last_name +
                               ' ' + u.first_name, u.surname])

            else:
                raise Http404

        elif state == 'participants':
            list_participants_weight = self.event.list_participants_weight()
            for e in list_participants_weight:
                u = e[0]
                ws.append([u.username, e[1], '', u.last_name +
                           ' ' + u.first_name, u.surname])

        elif state == 'registrants':
            list_registrants_weight = self.event.list_registrants_weight()
            for e in list_registrants_weight:
                u = e[0]
                ws.append([u.username, e[1], '', u.last_name +
                           ' ' + u.first_name, u.surname])
        else:
            raise Http404

        # Return the file
        response = HttpResponse(save_virtual_workbook(wb),
                   content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=event-'+str(self.event.datetime.date())+".xlsx"
        return response


class EventUploadXlsx(EventMixin, BorgiaFormView):
    """
    Upload Excel.
    """
    permission_required = 'events.change_event'
    menu_type = 'members'
    form_class = EventUploadXlsxForm
    allow_manager = True
    need_ongoing_event = True

    def form_valid(self, form):
        try:
            wb = load_workbook(self.request.FILES['list_user'], read_only=True)
            sheet = wb.active
            rows = sheet.rows
            next(rows)  # Skip the first row
        except:
            raise PermissionDenied

        if form.cleaned_data['state'] == 'participants':
            is_participant = True
        else:
            is_participant = False

        errors = []
        nb_empty_rows = 0
        min_col = sheet.min_column - 1
        i = 1

        # Enregistrement des pondérations
        for row in rows:
            i += 1
            try:
                if row[min_col].value and row[min_col + 1].value:
                    username = row[min_col].value.strip()  # Should be a str
                    if User.objects.filter(username=username).count() > 0:
                        user = User.objects.get(username=username)
                        try:
                            # Should be an int. Else, raise an error
                            pond = int(row[min_col+1].value)
                            if pond > 0:
                                self.event.change_weight(
                                    user, pond, is_participant)
                        except:
                            errors.append(
                                'Erreur avec ' + username + ' (ligne n*' + str(i) + '). A priori pas ajouté.')
                    else:
                        errors.append("L'utilisateur " + username +
                                      " n'existe pas. (ligne n*" + str(i) + ").")
                else:
                    nb_empty_rows += 1
            except:
                errors.append("Erreur avec la ligne n*" +
                              str(i) + ". Pas ajouté.")

        errors_count = len(errors)
        error_message = ""
        if errors_count >= 1:
            error_message = str(errors_count) + \
                " erreur(s) pendant l'ajout : \n - "
            if sheet.max_row - nb_empty_rows - 1 == errors_count:
                error_message += "Aucune donnée ne peut être importée (Vérifiez le format et la syntaxe du contenu du fichier)"
            else:
                error_message += "\n - ".join(errors)
            messages.warning(self.request, error_message)
            if i - nb_empty_rows - (1 + errors_count) > 0:
                messages.success(self.request, "Les " + str(i - nb_empty_rows -
                                                            (1 + errors_count)) + " autres utilisateurs ont bien été ajoutés.")
        else:
            messages.success(self.request, "Les " + str(i-1) +
                             " utilisateurs ont bien été ajoutés.")

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('url_event_manage_users',
                       kwargs={'pk': self.event.pk})
