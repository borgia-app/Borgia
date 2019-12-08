import datetime
import json
import random
import re
import string

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import (LoginRequiredMixin,
                                        PermissionRequiredMixin)
from django.contrib.auth.models import Group, Permission
from django.core.exceptions import ObjectDoesNotExist, PermissionDenied
from django.db.models import Q
from django.http import Http404, HttpResponseBadRequest
from django.shortcuts import HttpResponse, redirect, render
from django.urls import reverse
from django.utils.encoding import force_text

from borgia.utils import (get_members_group, human_unused_permissions,
                          get_permission_name_group_managing)
from borgia.views import BorgiaFormView, BorgiaView
from configurations.utils import configuration_get
from users.forms import (GroupUpdateForm, UserCreationCustomForm, UserDownloadXlsxForm,
                         UserSearchForm, UserUpdateForm, UserUploadXlsxForm)
from users.mixins import GroupMixin, UserMixin
from users.models import User


class UserListView(LoginRequiredMixin, PermissionRequiredMixin, BorgiaFormView):
    """
    List User instances.

    """
    permission_required = 'users.view_user'
    menu_type = 'managers'
    lm_active = 'lm_user_list'
    template_name = 'users/user_list.html'
    form_class = UserSearchForm

    search = None
    year = None
    state = None
    headers = {'username': 'asc',
               'last_name': 'asc',
               'surname': 'asc',
               'family': 'asc',
               'campus': 'asc',
               'year': 'asc',
               'balance': 'asc'}
    sort = None

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Header List
        context['list_header'] = [["username", "Username"], ["last_name", "Nom Prénom"], ["surname", "Bucque"], [
            "family", "Fam's"], ["campus", "Tabagn's"], ["year", "Prom's"], ["balance", "Solde"]]

        try:
            self.sort = self.request.GET['sort']
        except KeyError:
            pass

        query = User.objects.all()
        if self.sort is not None:
            context['sort'] = self.sort
            if self.headers[self.sort] == "des":
                context['reverse'] = True
                context['user_list'] = self.form_query(
                    query.order_by(self.sort).reverse())
                self.headers[self.sort] = "asc"
            else:
                context['user_list'] = self.form_query(
                    query.order_by(self.sort))
                self.headers[self.sort] = "des"
        else:
            context['user_list'] = self.form_query(
                query)

        return context

    def form_query(self, query):
        if self.search:
            query = query.filter(
                Q(last_name__icontains=self.search)
                | Q(first_name__icontains=self.search)
                | Q(surname__icontains=self.search)
                | Q(username__icontains=self.search)
            )

        if self.year and self.year != 'all':
            query = query.filter(
                year=self.year)

        if self.state and self.state != 'all':
            if self.state == 'internals':
                query = query.exclude(groups=get_members_group(
                    is_externals=True)).filter(is_active=True)
            if self.state == 'negative_balance':
                query = query.filter(balance__lt=0.0, is_active=True)
            elif self.state == 'threshold':
                threshold = configuration_get(
                    'BALANCE_THRESHOLD_PURCHASE').get_value()
                query = query.filter(balance__lt=threshold, is_active=True)
            elif self.state == 'unactive':
                query = query.filter(is_active=False)
        else:
            query = query.filter(is_active=True)

        return query

    def form_valid(self, form):
        if form.cleaned_data['search']:
            self.search = form.cleaned_data['search']

        if form.cleaned_data['state']:
            self.state = form.cleaned_data['state']

        if form.cleaned_data['year']:
            self.year = form.cleaned_data['year']

        return self.get(self.request, self.args, self.kwargs)

    def get_initial(self):
        initial = super().get_initial()
        initial['search'] = self.search
        return initial


class UserCreateView(LoginRequiredMixin, PermissionRequiredMixin, BorgiaFormView):
    """
    Create a new user and redirect to the workboard of the group.

    """
    permission_required = 'users.add_user'
    menu_type = 'managers'
    lm_active = 'lm_user_create'
    template_name = 'users/user_create.html'
    form_class = UserCreationCustomForm

    def __init__(self):
        self.object = None

    def form_valid(self, form):
        user = User.objects.create(username=form.cleaned_data['username'],
                                   first_name=form.cleaned_data['first_name'],
                                   last_name=form.cleaned_data['last_name'],
                                   email=form.cleaned_data['email'],
                                   surname=form.cleaned_data['surname'],
                                   family=form.cleaned_data['family'],
                                   campus=form.cleaned_data['campus'],
                                   year=form.cleaned_data['year'])
        user.set_password(form.cleaned_data['password'])
        user.save()

        is_external_member = form.cleaned_data['is_external_member']
        user.groups.add(get_members_group(is_external_member))

        user.save()

        # User object is assigned to self.object (so we can access to it in get_success_url)
        self.object = user

        return super().form_valid(form)

    def get_initial(self):
        initial = super().get_initial()
        initial['campus'] = 'Me'
        initial['year'] = datetime.datetime.now().year - 1
        return initial

    def get_success_message(self, cleaned_data):
        return "L'utilisateur a bien été crée'"

    def get_success_url(self):
        """
        If can retrieve user: go to the user.
        If not, go to the workboard of the group.
        """
        if self.request.user.has_perm('users.view_user'):
            return reverse('url_user_retrieve', kwargs={'user_pk': self.object.pk})
        else:
            return reverse('url_managers_workboard')


class UserRetrieveView(UserMixin, BorgiaView):
    """
    Retrieve a User instance.

    """
    permission_required = 'users.view_user'
    menu_type = "managers"
    template_name = 'users/user_retrieve.html'


    def get(self, request, *args, **kwargs):
        if self.user == self.request.user:
           self.menu_type = "members"
        elif self.user.has_perm('users.change_user'):
           self.menu_type = "managers"   
        context = self.get_context_data(**kwargs)
        return render(request, self.template_name, context=context)


class UserUpdateView(UserMixin, BorgiaFormView):
    """
    Update an user and redirect to the workboard of the group.

    """
    permission_required = 'users.change_user'
    menu_type = 'managers'
    template_name = 'users/user_update.html'
    form_class = UserUpdateForm
    model = User
    modified = False

    def has_permission(self):
        """
        Override has_permission to allow self update
        """
        self.is_manager = super().has_permission()
        self.is_self_update = self.user == self.request.user
        return self.is_manager or self.is_self_update

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.user
        kwargs['is_manager'] = self.is_manager
        kwargs['is_self_update'] = self.is_self_update
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        user = self.user
        for k in UserUpdateForm(user=user, is_manager=self.is_manager, is_self_update=self.is_self_update).fields.keys():
            initial[k] = getattr(user, k)
        if self.is_self_update:
           self.menu_type = "members"
        return initial

    def form_valid(self, form):
        user = self.user
        for k in form.fields.keys():
            if k == "avatar" and form.cleaned_data[k] is False:
                if user.avatar:
                    user.avatar.delete(True)
                    self.modified = True
            elif form.cleaned_data[k] != getattr(user, k):
                setattr(user, k, form.cleaned_data[k])
                self.modified = True
        user.save()

        return super().form_valid(form)

    def get_success_message(self, cleaned_data):
        if self.modified :
            return "Les informations ont bien été mises à jour"
        else:
            return "Pas de modification"

    def get_success_url(self):
        if self.is_manager:
            return reverse('url_user_retrieve',
                           kwargs={'user_pk': self.user.pk})
        else:
            return reverse('url_members_workboard')


class UserDeactivateView(UserMixin, BorgiaView):
    """
    Deactivate a user and redirect to the workboard of the group.

    """
    permission_required = 'users.delete_user'
    menu_type = 'managers'
    template_name = 'users/user_deactivate.html'
    success_message = "Le compte de %(user)s a bien été "
    error_event_message = "Veuillez attribuer la gestion des évènements suivants à un autre utilisateur avant de désactiver le compte:"

    def has_permission(self):
        has_perm = super().has_permission()
        if has_perm:
            return True
        else:
            return self.user == self.request.user

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        return render(request, self.template_name, context=context)

    def post(self, request, *args, **kwargs):
        deactivated = False
        if self.user.is_active is True:
            events = self.user.manager.filter(done=False)
            if events.count() > 0:
                for event in events:
                    self.error_event_message += "\n - " + event.description
                messages.warning(request, self.error_event_message)
            else:
                deactivated = True
                self.user.is_active = False
                # si c'est un gadz. Special members can't be added to other groups
                if get_members_group() in self.user.groups.all():
                    self.user.groups.set([get_members_group(), ])
                self.user.save()
        else:
            self.user.is_active = True
        self.user.save()

        if self.user.is_active:
            self.success_message += 'activé'
        else:
            self.success_message += 'désactivé'

        messages.success(request, self.success_message % dict(
            user=self.user,
        ))

        if request.user == self.user and deactivated:
            success_url = reverse(
                'url_logout')
        else:
            success_url = reverse(
                'url_user_retrieve',
                kwargs={'user_pk': self.user.pk})

        return redirect(force_text(success_url))


class GroupUpdateView(GroupMixin, BorgiaFormView):
    menu_type = 'managers'
    template_name = 'users/group_update.html'
    form_class = GroupUpdateForm

    def __init__(self):
        super().__init__()
        self.perm_manage_group = None

    def get_permission_required(self):
        """
        Override the permission_required attribute.
        Must return an iterable.
        """
        self.perm_manage_group = (
            get_permission_name_group_managing(self.group),)
        self.lm_active = 'lm_group_manage_' + self.group.name
        return self.perm_manage_group

    def get_form_kwargs(self):
        """
        Add possible members and permissions to kwargs of the form.

        Possible members are all members, except externals members and unactive users.
        Possible permissions are all permissions.
        :note:: For the special case of a shop management, two groups exist:
        group of chiefs and group of associates. If the group of associates is
        managed, possible permissions are only permissions of the chiefs group.
        """
        kwargs = super().get_form_kwargs()

        if self.group.name.startswith('associates-') is True:
            chiefs_group_name = self.group.name.replace(
                'associates', 'chiefs')
            query = Group.objects.get(name=chiefs_group_name).permissions.all().exclude(
                name=self.perm_manage_group[0])

        else:
            query = Permission.objects.all()

        kwargs['possible_permissions'] = query.exclude(
            pk__in=human_unused_permissions())
        kwargs['possible_members'] = User.objects.filter(is_active=True).exclude(
            groups=get_members_group(is_externals=True))
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        initial['members'] = self.group.user_set.all()
        initial['permissions'] = self.group.permissions.all()
        return initial

    def form_valid(self, form):
        """
        Update permissions and members of the group updated.
        """
        old_members = self.group.user_set.all()
        new_members = form.cleaned_data['members']
        old_permissions = self.group.permissions.all()
        new_permissions = form.cleaned_data['permissions']

        # Modification des membres
        for member in old_members:
            if member not in new_members:
                member.groups.remove(self.group)
                member.save()
        for member in new_members:
            if member not in old_members:
                member.groups.add(self.group)
                member.save()

        # Modification des permissions
        for perm in old_permissions:
            if perm not in new_permissions:
                self.group.permissions.remove(perm)
        for perm in new_permissions:
            if perm not in old_permissions:
                self.group.permissions.add(perm)
        self.group.save()

        return super().form_valid(form)

    def get_success_message(self, cleaned_data):
        return 'Le groupe ' + self.group.name + ' a bien été mis à jour'

    def get_success_url(self):
        return reverse('url_managers_workboard')


class UserUploadXlsxView(LoginRequiredMixin, PermissionRequiredMixin, BorgiaFormView):
    """
    Download/Upload Excel for adding users.
    """
    permission_required = 'users.add_user'
    menu_type = 'managers'
    template_name = 'users/bulk_edit.html'
    form_class = UserUploadXlsxForm
    second_form_class = UserDownloadXlsxForm
    lm_active = 'lm_user_create'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'upload_form' not in context:
            context['upload_form'] = self.form_class()
        if 'download_form' not in context:
            context['download_form'] = self.second_form_class()
        return context

    def form_valid(self, form):
        try:
            wb = openpyxl.load_workbook(
                self.request.FILES['list_user'], read_only=True)
            sheet = wb.active
            rows = sheet.rows
        except:
            raise PermissionDenied

        errors = []
        nb_empty_rows = 0
        min_col = sheet.min_column
        max_col = sheet.max_column
        min_row = sheet.min_row
        i = min_row

        columns = form.cleaned_data['xlsx_columns']
        # Setting column numbers
        for col in range(min_col, max_col+1):
            if sheet.cell(None, sheet.min_row, col).value == 'username':
                col_username = col - min_row
            elif sheet.cell(None, sheet.min_row, col).value == 'first_name':
                col_first_name = col - min_row
            elif sheet.cell(None, sheet.min_row, col).value == 'last_name':
                col_last_name = col - min_row
            elif sheet.cell(None, sheet.min_row, col).value == 'email':
                col_email = col - min_row
            elif sheet.cell(None, sheet.min_row, col).value == 'surname':
                col_surname = col - min_row
            elif sheet.cell(None, sheet.min_row, col).value == 'family':
                col_family = col - min_row
            elif sheet.cell(None, sheet.min_row, col).value == 'campus':
                col_campus = col - min_row
            elif sheet.cell(None, sheet.min_row, col).value == 'year':
                col_year = col - min_row

        for _ in range(min_row):
            next(rows)

        skipped_rows = []
        # Saving users
        for row in rows:
            i += 1
            try:
                user_dict = {}
                errors_on_required_columns = []
                skipped_row = False
                try:
                    if row[col_username].value:
                        user_dict['username'] = row[col_username].value.strip()
                    else:
                        skipped_row = True
                        nb_empty_rows += 1
                except:
                    skipped_row = True
                    errors_on_required_columns.append('username')

                if 'first_name' in columns:
                    try:
                        if row[col_first_name].value:
                            user_dict['first_name'] = row[col_first_name].value.strip()
                    except:
                        errors_on_required_columns.append('first_name')

                if 'last_name' in columns:
                    try:
                        if row[col_last_name].value:
                            user_dict['last_name'] = row[col_last_name].value.strip()
                    except:
                        errors_on_required_columns.append('last_name')

                if 'email' in columns:
                    try:
                        if row[col_email].value:
                            user_dict['email'] = row[col_email].value.strip()
                    except:
                        errors_on_required_columns.append('email')

                if 'surname' in columns:
                    try:
                        if row[col_surname].value:
                            user_dict['surname'] = row[col_surname].value.strip()
                    except:
                        errors_on_required_columns.append('surname')

                if 'family' in columns:
                    try:
                        if row[col_family].value:
                            user_dict['family'] = str(
                                row[col_family].value).strip()
                    except:
                        errors_on_required_columns.append('family')

                if 'campus' in columns:
                    try:
                        if row[col_campus].value:
                            user_dict['campus'] = row[col_campus].value.strip()
                    except:
                        errors_on_required_columns.append('campus')

                if 'year' in columns:
                    try:
                        if row[col_year].value:
                            user_dict['year'] = int(row[col_year].value)
                    except:
                        errors_on_required_columns.append('year')

                if not skipped_row:
                    username = user_dict['username']
                    if User.objects.filter(username=username).count() > 0:
                        User.objects.filter(
                            username=username).update(**user_dict)
                    else:
                        user = User.objects.create(**user_dict)
                        user.set_password(''.join(random.SystemRandom().choice(string.ascii_uppercase + string.digits)
                                                  for _ in range(20)))
                    user = User.objects.get(username=username)
                    user.save()

                    user.groups.add(get_members_group())
                else:
                    skipped_rows.append(str(i))
            except:
                if str(i) in skipped_rows:
                    errors.append("La ligne n*" + str(i) +
                                  "n'a pas été traitée car le username est manquant")  # TODO: Make this error appear
                elif errors_on_required_columns:
                    errors.append("Les colonnes " + ", ".join(errors_on_required_columns) +
                                  " sont requis et comportent des erreurs (ligne n*" + str(i) + ")")
                else:
                    errors.append(
                        "Une erreur empêche le traitement du fichier Excel (ligne n*" + str(i) + ")")

        error_message = ""
        error_message += "\n - ".join(errors)
        messages.success(self.request, str(i - len(skipped_rows) - min_row - len(errors)) +
                         " utilisateurs ont été crées/mis à jour")
        messages.warning(self.request, error_message)

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('url_user_list')


class UserAddByListXlsxDownload(LoginRequiredMixin, PermissionRequiredMixin, BorgiaView):
    """
    Download Excel for adding users.
    """
    permission_required = 'users.add_user'
    menu_type = 'managers'
    lm_active = 'lm_user_create'

    def get(self, request, *args, **kwargs):
        columns = request.GET.getlist('xlsx_columns')

        if columns:
            columns = ['first_name', 'last_name', 'email', 'surname',
                       'family', 'campus', 'year']

        columns.insert(0, 'username')

        wb = openpyxl.Workbook()
        # grab the active worksheet
        ws = wb.active
        ws.title = "users"
        ws.append(columns)
        for col in ['A','B','C','D','E']:
            ws.column_dimensions[col].width = 30

        # Return the file
        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="UsersList.xlsx"'

        users = User.objects.all().values_list(*columns)
        for user in users:
            ws.append(user)
        return response


def username_from_username_part(request):
    data = []
    try:
        key = request.GET.get('keywords')
        regex = r"^" + re.escape(key) + r"(\W|$)"

        where_search = User.objects.filter(
            family__regex=regex, is_active=True).order_by('-year')
        if len(key) > 2:
            if key.isalpha():
                # Nom de famille, début ou entier à partir de 3 caractères
                where_search = where_search | User.objects.filter(
                    last_name__istartswith=key, is_active=True)
                # Prénom, début ou entier à partir de 3 caractères
                where_search = where_search | User.objects.filter(
                    first_name__istartswith=key, is_active=True)
                # Buque, début ou entier à partir de 3 caractères
                where_search = where_search | User.objects.filter(
                    surname__istartswith=key, is_active=True)

                # Suppression des doublons
                where_search = where_search.distinct()

        for e in where_search:
            data.append(e.username)
    except KeyError:
        pass

    return HttpResponse(json.dumps(data))


@login_required
def balance_from_username(request):
    operator = request.user

    if operator.has_perm('modules.use_operatorsalemodule'):
        try:
            username = request.GET['username']
            data = str(User.objects.get(username=username).balance)
            return HttpResponse(json.dumps(data))
        except KeyError:
            return HttpResponseBadRequest()
        except ObjectDoesNotExist:
            return HttpResponseBadRequest()
    else:
        raise PermissionDenied
